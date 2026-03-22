"""Excel export handler with formatting."""

import pandas as pd
from typing import List
from models.conference import Conference
from utils.logger import get_logger
from datetime import date, timedelta
import os

logger = get_logger(__name__)


class ExcelHandler:
    """Handle exporting conference data to Excel with formatting."""

    def __init__(self, output_path: str = "data/processed/conferences.xlsx"):
        """
        Initialize Excel handler.

        Args:
            output_path: Path to Excel file
        """
        self.output_path = output_path

        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

    def export_conferences(self, conferences: List[Conference]) -> bool:
        """
        Export conferences to Excel with formatting.

        Args:
            conferences: List of Conference objects

        Returns:
            True if successful
        """
        if not conferences:
            logger.warning("No conferences to export")
            return False

        try:
            # Convert to DataFrame
            data = [conf.to_dict() for conf in conferences]
            df = pd.DataFrame(data)

            # Create Excel writer
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Main sheet
                df.to_excel(writer, sheet_name='All Conferences', index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['All Conferences']

                # Format headers
                self._format_headers(worksheet)

                # Add hyperlinks to website_url and cfp_url columns
                self._add_hyperlinks(worksheet, df)

                # Add conditional formatting for deadlines
                self._add_deadline_formatting(worksheet, df)

                # Auto-adjust column widths
                self._auto_adjust_columns(worksheet)

                # Create additional sheets by year
                self._create_year_sheets(writer, conferences)

            logger.info(f"Exported {len(conferences)} conferences to {self.output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            return False

    def _format_headers(self, worksheet):
        """Format header row."""
        from openpyxl.styles import Font, PatternFill, Alignment

        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _add_hyperlinks(self, worksheet, df):
        """Add hyperlinks to URL columns."""
        from openpyxl.styles import Font

        # Find URL columns
        url_columns = []
        for idx, col in enumerate(df.columns):
            if 'url' in col.lower():
                url_columns.append((idx + 1, col))  # +1 because Excel is 1-indexed

        # Add hyperlinks
        for col_idx, col_name in url_columns:
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

    def _add_deadline_formatting(self, worksheet, df):
        """Add conditional formatting to highlight upcoming deadlines."""
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import CellIsRule

        # Find deadline columns
        deadline_col = None
        for idx, col in enumerate(df.columns):
            if 'submission_deadline' in col.lower():
                deadline_col = idx + 1
                break

        if not deadline_col:
            return

        # Highlight deadlines in next 30 days (yellow)
        today = date.today()
        warning_date = today + timedelta(days=30)

        # Apply formatting to each cell
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=deadline_col)
            if cell.value:
                try:
                    # Parse date
                    from utils.date_parser import parse_date
                    deadline = parse_date(str(cell.value))

                    if deadline:
                        # Upcoming deadline (yellow)
                        if today <= deadline <= warning_date:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        # Very soon (< 7 days) - red
                        elif today <= deadline <= today + timedelta(days=7):
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            from openpyxl.styles import Font
                            cell.font = Font(color="FFFFFF", bold=True)

                        # Past deadline (gray)
                        elif deadline < today:
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                except:
                    pass

    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _create_year_sheets(self, writer, conferences: List[Conference]):
        """Create separate sheets for each year."""
        # Group by year
        by_year = {}
        for conf in conferences:
            if conf.year:
                if conf.year not in by_year:
                    by_year[conf.year] = []
                by_year[conf.year].append(conf)

        # Create sheet for each year
        for year, year_confs in sorted(by_year.items()):
            data = [conf.to_dict() for conf in year_confs]
            df = pd.DataFrame(data)

            sheet_name = f"Year {year}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Format the sheet
            worksheet = writer.sheets[sheet_name]
            self._format_headers(worksheet)
            self._add_hyperlinks(worksheet, df)
            self._add_deadline_formatting(worksheet, df)
            self._auto_adjust_columns(worksheet)

        logger.info(f"Created {len(by_year)} year-based sheets")
